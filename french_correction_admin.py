import time
import re
from io import BytesIO

import pandas as pd
import streamlit as st
import language_tool_python
from openpyxl import load_workbook
from openpyxl.styles import Alignment


SHEET_NAME = "All Answers"
LANGUAGE = "fr-FR"

ORIGINAL_COL_INDEX = 5  # Column F

SUGGESTED_COL_NAME = "Suggested_Correction"  # Column G
DECISION_COL_NAME = "Admin_Decision"         # Column H
FINAL_COL_NAME = "Final_Answer"              # Column I
NOTES_COL_NAME = "Correction_Notes"          # Column J

PROTECTED_TERMS = [
    "USJ",
    "Université Saint-Joseph",
    "Université Saint Joseph",
    "Saint-Joseph",
    "Saint Joseph"
]


st.set_page_config(
    page_title="Correction orthographique - Français France",
    layout="wide"
)

st.title("Correction orthographique - Français France")

st.caption(
    "Upload the Excel file. The app previews the full sheet first, then checks Column F progressively and adds corrections in Columns G to J."
)


@st.cache_resource
def load_tool():
    return language_tool_python.LanguageTool(LANGUAGE)


def get_error_length(match):
    return getattr(match, "errorLength", getattr(match, "error_length", 0))


def match_overlaps_protected_term(text, match, protected_terms):
    error_length = get_error_length(match)
    start = match.offset
    end = match.offset + error_length

    for term in protected_terms:
        for found in re.finditer(re.escape(term), text, flags=re.IGNORECASE):
            term_start, term_end = found.span()

            if start < term_end and end > term_start:
                return True

    return False


def correct_french_text(text, tool, protected_terms):
    if pd.isna(text) or str(text).strip() == "":
        return "", ""

    text = str(text)

    matches = tool.check(text)

    filtered_matches = [
        match for match in matches
        if not match_overlaps_protected_term(text, match, protected_terms)
    ]

    corrected = language_tool_python.utils.correct(text, filtered_matches)

    notes = []

    for match in filtered_matches:
        if match.replacements:
            error_length = get_error_length(match)
            error_text = text[match.offset:match.offset + error_length]
            suggestions = ", ".join(match.replacements[:5])

            notes.append(
                f"{error_text} → {suggestions} ({match.message})"
            )

    return corrected, " | ".join(notes)


def prepare_dataframe(df):
    if df.shape[1] < 6:
        st.error("Column F was not found. The sheet must contain at least 6 columns.")
        st.stop()

    answer_column = df.columns[ORIGINAL_COL_INDEX]

    if SUGGESTED_COL_NAME not in df.columns:
        df.insert(6, SUGGESTED_COL_NAME, "")

    if DECISION_COL_NAME not in df.columns:
        df.insert(7, DECISION_COL_NAME, "Pending")

    if FINAL_COL_NAME not in df.columns:
        df.insert(8, FINAL_COL_NAME, df[answer_column])

    if NOTES_COL_NAME not in df.columns:
        df.insert(9, NOTES_COL_NAME, "")

    return df, answer_column


def dataframe_to_excel_preserve_workbook(original_file_bytes, df):
    input_stream = BytesIO(original_file_bytes)
    workbook = load_workbook(input_stream)

    if SHEET_NAME not in workbook.sheetnames:
        st.error(f'Sheet "{SHEET_NAME}" was not found in the workbook.')
        st.stop()

    worksheet = workbook[SHEET_NAME]

    headers = list(df.columns)

    for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_index).value = header

    for row_index in range(len(df)):
        excel_row = row_index + 2

        for col_index, col_name in enumerate(headers, start=1):
            value = df.iloc[row_index][col_name]

            if pd.isna(value):
                value = ""

            worksheet.cell(row=excel_row, column=col_index).value = value

    worksheet.freeze_panes = "A2"

    for col_cells in worksheet.columns:
        column_letter = col_cells[0].column_letter

        if column_letter in ["F", "G", "I", "J"]:
            worksheet.column_dimensions[column_letter].width = 70
        elif column_letter == "H":
            worksheet.column_dimensions[column_letter].width = 25
        else:
            worksheet.column_dimensions[column_letter].width = 20

        for cell in col_cells:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    output = BytesIO()
    workbook.save(output)

    return output.getvalue()


def reset_session_for_new_file(uploaded_file, file_bytes):
    file_signature = f"{uploaded_file.name}_{uploaded_file.size}"

    if st.session_state.get("file_signature") != file_signature:
        for key in list(st.session_state.keys()):
            if (
                key.startswith("decision_")
                or key.startswith("manual_")
                or key.startswith("orig_")
            ):
                del st.session_state[key]

        for key in [
            "file_signature",
            "original_file_bytes",
            "df",
            "df_raw",
            "answer_column",
            "processed_rows",
            "current_row"
        ]:
            if key in st.session_state:
                del st.session_state[key]

        st.session_state["file_signature"] = file_signature
        st.session_state["original_file_bytes"] = file_bytes


uploaded_file = st.file_uploader(
    "Upload Excel file",
    type=["xlsx"],
    help="Upload the Excel file containing the sheet: All Answers"
)

if uploaded_file is None:
    st.info("Please upload the Excel file to start.")
    st.stop()

file_bytes = uploaded_file.getvalue()

reset_session_for_new_file(uploaded_file, file_bytes)

if "df" not in st.session_state:
    df_raw = pd.read_excel(BytesIO(file_bytes), sheet_name=SHEET_NAME)
    df_prepared, answer_column = prepare_dataframe(df_raw.copy())

    st.session_state["df_raw"] = df_raw
    st.session_state["df"] = df_prepared
    st.session_state["answer_column"] = answer_column
    st.session_state["processed_rows"] = set()
    st.session_state["current_row"] = 0

df_raw = st.session_state["df_raw"]
df = st.session_state["df"]
answer_column = st.session_state["answer_column"]

st.success(f"Loaded sheet: {SHEET_NAME}")
st.info(f"Answers detected in Column F: {answer_column}")

st.markdown("## Full Excel preview before correction")
st.caption("This is the uploaded Excel sheet before any orthographic correction.")

st.dataframe(
    df_raw,
    use_container_width=True,
    height=450
)

with st.expander("Protected terms not considered as spelling mistakes"):
    protected_terms_text = st.text_area(
        "Protected terms",
        value="\n".join(PROTECTED_TERMS),
        height=140,
        help="Each term on a separate line. These terms will be ignored by LanguageTool."
    )

protected_terms = [
    term.strip()
    for term in protected_terms_text.splitlines()
    if term.strip()
]

tool = load_tool()

total_rows = len(df)

non_empty_rows = [
    i for i in range(total_rows)
    if str(df.loc[i, answer_column]).strip() not in ["", "nan", "None"]
]

st.markdown("## Row-by-row correction")

col_a, col_b, col_c = st.columns(3)

with col_a:
    st.metric("Total rows", total_rows)

with col_b:
    st.metric("Non-empty answers", len(non_empty_rows))

with col_c:
    st.metric("Processed rows", len(st.session_state["processed_rows"]))


st.markdown("### Processing controls")

batch_size = st.selectbox(
    "Number of rows to check per click",
    options=[10, 25, 50],
    index=0
)

process_next = st.button("Check next rows", type="primary")
process_all = st.button("Check all remaining rows")

progress = st.progress(
    len(st.session_state["processed_rows"]) / max(len(non_empty_rows), 1)
)

status_box = st.empty()

if process_next or process_all:
    rows_to_process = []

    for i in non_empty_rows:
        if i not in st.session_state["processed_rows"]:
            rows_to_process.append(i)

    if not process_all:
        rows_to_process = rows_to_process[:batch_size]

    if not rows_to_process:
        st.success("All rows are already checked.")
    else:
        for counter, i in enumerate(rows_to_process, start=1):
            original_text = str(df.loc[i, answer_column])

            global_checked = len(st.session_state["processed_rows"]) + 1

            status_box.write(
                f"Checking answer {global_checked} of {len(non_empty_rows)} "
                f"- Excel row {i + 2}"
            )

            corrected_text, correction_notes = correct_french_text(
                original_text,
                tool,
                protected_terms
            )

            df.loc[i, SUGGESTED_COL_NAME] = corrected_text
            df.loc[i, NOTES_COL_NAME] = correction_notes

            if str(corrected_text).strip() == str(original_text).strip():
                df.loc[i, DECISION_COL_NAME] = "No correction"
                df.loc[i, FINAL_COL_NAME] = original_text
            else:
                df.loc[i, DECISION_COL_NAME] = "Pending"
                df.loc[i, FINAL_COL_NAME] = original_text

            st.session_state["processed_rows"].add(i)

            progress.progress(
                len(st.session_state["processed_rows"]) / max(len(non_empty_rows), 1)
            )

            time.sleep(0.05)

        st.session_state["df"] = df
        status_box.success("Processing completed for selected rows.")


st.markdown("## Admin validation")

show_only_pending = st.checkbox(
    "Show only rows with suggested corrections pending validation",
    value=True
)

processed_indices = sorted(list(st.session_state["processed_rows"]))

if show_only_pending:
    display_indices = [
        i for i in processed_indices
        if df.loc[i, DECISION_COL_NAME] == "Pending"
    ]
else:
    display_indices = processed_indices

if not display_indices:
    st.info("No rows to validate yet. Click 'Check next rows' or 'Check all remaining rows'.")
else:
    for i in display_indices:
        original = str(df.loc[i, answer_column] or "")
        suggested = str(df.loc[i, SUGGESTED_COL_NAME] or "")
        notes = str(df.loc[i, NOTES_COL_NAME] or "")

        with st.expander(f"Excel row {i + 2}", expanded=False):
            meta_cols = st.columns(5)

            if "Respondent_Type" in df.columns:
                meta_cols[0].markdown(f"**Type:** {df.loc[i, 'Respondent_Type']}")

            if "groupe" in df.columns:
                meta_cols[1].markdown(f"**Groupe:** {df.loc[i, 'groupe']}")

            if "section" in df.columns:
                meta_cols[2].markdown(f"**Section:** {df.loc[i, 'section']}")

            if "category" in df.columns:
                meta_cols[3].markdown(f"**Category:** {df.loc[i, 'category']}")

            meta_cols[4].markdown(f"**Row:** {i + 2}")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("#### Original answer - Column F")
                st.text_area(
                    "Original answer",
                    value=original,
                    height=220,
                    disabled=True,
                    key=f"orig_{i}",
                    label_visibility="collapsed"
                )

            with col2:
                st.markdown("#### Suggested correction - Column G")
                manual_correction = st.text_area(
                    "Suggested correction",
                    value=suggested,
                    height=220,
                    key=f"manual_{i}",
                    label_visibility="collapsed"
                )

            st.markdown("#### Detected issues")

            if notes.strip():
                st.caption(notes)
            else:
                st.caption("No issue detected.")

            decision_options = [
                "Pending",
                "Accept correction",
                "Reject correction"
            ]

            current_decision = df.loc[i, DECISION_COL_NAME]

            if current_decision not in decision_options:
                current_decision = "Pending"

            decision = st.radio(
                "Admin decision",
                decision_options,
                horizontal=True,
                key=f"decision_{i}",
                index=decision_options.index(current_decision)
            )

            df.loc[i, DECISION_COL_NAME] = decision
            df.loc[i, SUGGESTED_COL_NAME] = manual_correction

            if decision == "Accept correction":
                df.loc[i, FINAL_COL_NAME] = manual_correction

            elif decision == "Reject correction":
                df.loc[i, FINAL_COL_NAME] = original

            else:
                df.loc[i, FINAL_COL_NAME] = original

            st.session_state["df"] = df


st.markdown("## Full modified Excel preview")
st.caption("This preview includes all original columns plus Columns G to J.")

st.dataframe(
    df,
    use_container_width=True,
    height=550
)

excel_bytes = dataframe_to_excel_preserve_workbook(
    st.session_state["original_file_bytes"],
    df
)

st.download_button(
    "Download modified Excel with corrections",
    data=excel_bytes,
    file_name="FGs_All_Answers_With_Corrections.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
